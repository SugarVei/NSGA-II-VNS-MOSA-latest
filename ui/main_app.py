"""
多窗口应用主入口
Multi-Window Application Entry Point

实现三窗口流程: 主窗口(参数输入) -> 数据编辑窗口 -> 矩阵编码窗口 -> 结果展示窗口
"""

import sys
import os

# 设置Qt平台插件路径 (解决 "Could not find the Qt platform plugin" 错误)
if hasattr(sys, 'frozen'):
    plugin_path = os.path.join(os.path.dirname(sys.executable), 'platforms')
else:
    import PyQt5
    plugin_path = os.path.join(os.path.dirname(PyQt5.__file__), 'Qt5', 'plugins', 'platforms')
    if not os.path.exists(plugin_path):
        plugin_path = os.path.join(os.path.dirname(PyQt5.__file__), 'Qt', 'plugins', 'platforms')
if os.path.exists(plugin_path):
    os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path

import numpy as np

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QGroupBox, QLabel, QLineEdit, QSpinBox, QDoubleSpinBox,
    QPushButton, QRadioButton, QButtonGroup, QComboBox, QMessageBox,
    QFrame, QScrollArea, QSplitter, QTabWidget, QTextEdit, QTableWidget,
    QTableWidgetItem, QHeaderView, QDialog, QDialogButtonBox
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QPalette, QColor

from models.problem import SchedulingProblem
from models.solution import Solution


# ==================== 样式常量 ====================
# 现代化配色方案 - 温暖米色系
COLORS = {
    'background': '#E8E4DF',      # 温暖米色背景
    'card': '#FFFFFF',            # 卡片白色
    'dark_card': '#2D3436',       # 深色卡片
    'primary': '#F5A623',         # 温暖橙黄色主色
    'primary_hover': '#E09612',   # 深橙色（悬停）
    'secondary': '#3498db',       # 蓝色次要色
    'success': '#27ae60',         # 成功绿色
    'warning': '#f39c12',         # 警告橙色
    'danger': '#e74c3c',          # 危险红色
    'text_dark': '#2D3436',       # 深色文字
    'text_light': '#FFFFFF',      # 浅色文字
    'text_muted': '#7f8c8d',      # 柔和文字
    'border': '#D5D0CA',          # 边框色
    'border_focus': '#F5A623',    # 聚焦边框色
}

MAIN_STYLE = f"""
QMainWindow, QDialog, QWidget {{
    background-color: {COLORS['background']};
    font-family: "Microsoft YaHei", "Segoe UI", sans-serif;
}}
QScrollArea {{
    background-color: {COLORS['background']};
    border: none;
}}
QScrollArea > QWidget > QWidget {{
    background-color: {COLORS['background']};
}}
QGroupBox {{
    font-weight: bold;
    font-size: 16px;
    border: none;
    border-radius: 16px;
    margin-top: 24px;
    padding: 24px 20px 20px 20px;
    background-color: {COLORS['card']};
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 20px;
    padding: 6px 16px;
    color: {COLORS['text_dark']};
    font-size: 16px;
    font-weight: bold;
    background-color: {COLORS['card']};
    border-radius: 8px;
}}
QPushButton {{
    background-color: {COLORS['primary']};
    color: {COLORS['text_light']};
    border: none;
    padding: 12px 28px;
    border-radius: 12px;
    font-size: 14px;
    font-weight: bold;
    min-height: 20px;
}}
QPushButton:hover {{
    background-color: {COLORS['primary_hover']};
}}
QPushButton:pressed {{
    background-color: #C47F0A;
}}
QPushButton:disabled {{
    background-color: #D5D0CA;
    color: #A0A0A0;
}}
QSpinBox, QDoubleSpinBox, QLineEdit, QComboBox {{
    padding: 10px 14px;
    border: 2px solid {COLORS['border']};
    border-radius: 10px;
    background-color: {COLORS['card']};
    min-width: 100px;
    min-height: 20px;
    font-size: 14px;
    font-weight: bold;
    color: {COLORS['text_dark']};
}}
QSpinBox:focus, QDoubleSpinBox:focus, QLineEdit:focus, QComboBox:focus {{
    border: 2px solid {COLORS['border_focus']};
    background-color: #FFFDF8;
}}
QSpinBox::up-button, QSpinBox::down-button,
QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {{
    width: 24px;
    border: none;
    background-color: transparent;
}}
QLabel {{
    color: {COLORS['text_dark']};
    font-size: 14px;
    font-weight: bold;
    padding: 6px 8px;
    background-color: transparent;
}}
QRadioButton {{
    font-size: 16px;
    font-weight: bold;
    color: {COLORS['text_dark']};
    padding: 6px;
    background-color: transparent;
}}
QRadioButton::indicator {{
    width: 20px;
    height: 20px;
}}
QRadioButton::indicator:checked {{
    background-color: {COLORS['primary']};
    border: 2px solid {COLORS['primary']};
    border-radius: 10px;
}}
QRadioButton::indicator:unchecked {{
    background-color: {COLORS['card']};
    border: 2px solid {COLORS['border']};
    border-radius: 10px;
}}
QTableWidget {{
    gridline-color: {COLORS['border']};
    background-color: {COLORS['card']};
    border-radius: 12px;
    font-size: 13px;
    border: 1px solid {COLORS['border']};
    selection-background-color: #FFF3CD;
}}
QTableWidget::item {{
    padding: 8px;
    min-height: 36px;
}}
QHeaderView::section {{
    background-color: {COLORS['dark_card']};
    color: {COLORS['text_light']};
    padding: 10px 8px;
    font-weight: bold;
    font-size: 13px;
    border: none;
    min-height: 40px;
}}
QTabWidget::pane {{
    border: none;
    border-radius: 12px;
    background-color: {COLORS['card']};
    padding: 10px;
}}
QTabBar::tab {{
    padding: 10px 20px;
    margin-right: 4px;
    border: none;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
    background-color: {COLORS['border']};
    color: {COLORS['text_dark']};
    font-weight: bold;
}}
QTabBar::tab:selected {{
    background-color: {COLORS['card']};
    color: {COLORS['primary']};
}}
QTabBar::tab:hover {{
    background-color: #F5F5F5;
}}
QComboBox::drop-down {{
    border: none;
    width: 28px;
}}
QComboBox::down-arrow {{
    image: none;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 6px solid {COLORS['text_muted']};
}}
QComboBox QAbstractItemView {{
    background-color: {COLORS['card']};
    border: 1px solid {COLORS['border']};
    border-radius: 8px;
    selection-background-color: {COLORS['primary']};
    selection-color: {COLORS['text_light']};
}}
"""


PRIMARY_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {COLORS['success']};
    color: {COLORS['text_light']};
    border: none;
    padding: 14px 32px;
    border-radius: 12px;
    font-size: 15px;
    font-weight: bold;
}}
QPushButton:hover {{
    background-color: #219a52;
}}
QPushButton:pressed {{
    background-color: #1e8449;
}}
"""

SECONDARY_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {COLORS['dark_card']};
    color: {COLORS['text_light']};
    border: none;
    padding: 14px 32px;
    border-radius: 12px;
    font-size: 15px;
    font-weight: bold;
}}
QPushButton:hover {{
    background-color: #3d4548;
}}
QPushButton:pressed {{
    background-color: #1d2426;
}}
"""


class DataEditorDialog(QDialog):
    """
    数据编辑对话框 - 显示和编辑生成的或手动输入的数据
    """
    
    def __init__(self, parent=None, problem: SchedulingProblem = None, 
                 is_manual: bool = False, params: dict = None):
        super().__init__(parent)
        self.problem = problem
        self.is_manual = is_manual
        self.params = params or {}
        self.setup_ui()
        self.setStyleSheet(MAIN_STYLE)
        
        if problem:
            self.load_problem_data()
    
    def setup_ui(self):
        """初始化UI"""
        title = "手动输入数据" if self.is_manual else "查看/编辑生成的数据"
        self.setWindowTitle(title)
        self.setMinimumSize(1000, 700)
        
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)
        
        # 标签页
        self.tabs = QTabWidget()
        
        # 1. 加工时间
        self.create_processing_time_tab()
        
        # 2. 换模时间
        self.create_setup_time_tab()
        
        # 3. 能耗参数
        self.create_energy_tab()
        
        # 4. 工人参数
        self.create_worker_tab()
        
        main_layout.addWidget(self.tabs)
        
        # 按钮
        btn_layout = QHBoxLayout()
        
        # 复制当前表格按钮
        copy_btn = QPushButton("📋 复制当前表格")
        copy_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        copy_btn.clicked.connect(self.copy_current_table)
        btn_layout.addWidget(copy_btn)
        
        # 导出Excel按钮
        export_btn = QPushButton("📊 导出Excel")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        export_btn.clicked.connect(self.export_to_excel)
        btn_layout.addWidget(export_btn)
        
        if self.is_manual:
            random_btn = QPushButton("🎲 随机填充")
            random_btn.clicked.connect(self.random_fill)
            btn_layout.addWidget(random_btn)
        
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        confirm_btn = QPushButton("✅ 确认数据")
        confirm_btn.setStyleSheet(PRIMARY_BUTTON_STYLE)
        confirm_btn.clicked.connect(self.accept)
        btn_layout.addWidget(confirm_btn)
        
        main_layout.addLayout(btn_layout)
    
    def copy_current_table(self):
        """复制当前显示的表格到剪贴板"""
        from PyQt5.QtWidgets import QApplication
        
        # 获取当前标签页的表格
        current_index = self.tabs.currentIndex()
        table = None
        table_name = ""
        
        if current_index == 0:  # 加工时间
            table = self.proc_table
            table_name = "加工时间"
        elif current_index == 1:  # 换模时间
            table = self.setup_table
            table_name = "换模时间"
        elif current_index == 2:  # 能耗参数
            table = self.energy_table
            table_name = "能耗参数"
        elif current_index == 3:  # 工人参数
            table = self.worker_table
            table_name = "工人参数"
        
        if table is None:
            return
        
        # 构建表格文本
        rows = table.rowCount()
        cols = table.columnCount()
        
        # 添加表头
        header_row = []
        for col in range(cols):
            header = table.horizontalHeaderItem(col)
            header_row.append(header.text() if header else "")
        text = "\t".join(header_row) + "\n"
        
        # 添加数据行
        for row in range(rows):
            row_data = []
            # 添加行头
            v_header = table.verticalHeaderItem(row)
            if v_header:
                row_data.append(v_header.text())
            # 添加单元格数据
            for col in range(cols):
                item = table.item(row, col)
                row_data.append(item.text() if item else "")
            text += "\t".join(row_data) + "\n"
        
        # 复制到剪贴板
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        
        QMessageBox.information(self, "复制成功", f"已复制【{table_name}】表格到剪贴板！\n可粘贴到Excel或其他应用中。")
    
    def export_to_excel(self):
        """导出所有数据到Excel文件"""
        from PyQt5.QtWidgets import QFileDialog
        import os
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.warning(self, "缺少依赖", "请先安装openpyxl库:\npip install openpyxl")
            return
        
        # 选择保存路径
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", "调度问题数据.xlsx", "Excel文件 (*.xlsx)"
        )
        if not file_path:
            return
        
        p = self.problem
        wb = openpyxl.Workbook()
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # ===== 1. 加工时间 Sheet =====
        ws_proc = wb.active
        ws_proc.title = "加工时间"
        row = 1
        
        for stage in range(p.n_stages):
            for machine in range(p.machines_per_stage[stage]):
                # 标题行
                ws_proc.cell(row, 1, f"阶段{stage+1} - M{stage+1},{machine+1}").font = title_font
                row += 1
                
                # 表头
                headers = ["工件"] + [f"速度{s+1}" for s in range(p.n_speed_levels)]
                for col, h in enumerate(headers, 1):
                    cell = ws_proc.cell(row, col, h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                row += 1
                
                # 数据行
                for job in range(p.n_jobs):
                    ws_proc.cell(row, 1, f"工件{job+1}").border = thin_border
                    for speed in range(p.n_speed_levels):
                        val = p.processing_time[job, stage, machine, speed]
                        cell = ws_proc.cell(row, speed + 2, float(val))
                        cell.border = thin_border
                    row += 1
                row += 1  # 空行分隔
        
        # ===== 2. 换模时间 Sheet =====
        ws_setup = wb.create_sheet("换模时间")
        row = 1
        
        if p.setup_time is not None:
            for stage in range(p.n_stages):
                for machine in range(p.machines_per_stage[stage]):
                    ws_setup.cell(row, 1, f"阶段{stage+1} - M{stage+1},{machine+1}").font = title_font
                    row += 1
                    
                    # 表头
                    headers = ["前序\\后序"] + [f"→工件{j+1}" for j in range(p.n_jobs)]
                    for col, h in enumerate(headers, 1):
                        cell = ws_setup.cell(row, col, h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = thin_border
                    row += 1
                    
                    # 数据行
                    for j1 in range(p.n_jobs):
                        ws_setup.cell(row, 1, f"工件{j1+1}→").border = thin_border
                        for j2 in range(p.n_jobs):
                            val = p.setup_time[stage, machine, j1, j2]
                            cell = ws_setup.cell(row, j2 + 2, float(val))
                            cell.border = thin_border
                        row += 1
                    row += 1
        
        # ===== 3. 能耗参数 Sheet =====
        ws_energy = wb.create_sheet("能耗参数")
        row = 1
        
        for stage in range(p.n_stages):
            ws_energy.cell(row, 1, f"阶段{stage+1}").font = title_font
            row += 1
            
            # 表头
            headers = ["机器"] + [f"加工(速度{s+1})" for s in range(p.n_speed_levels)] + ["换模功率", "空闲功率"]
            for col, h in enumerate(headers, 1):
                cell = ws_energy.cell(row, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row += 1
            
            # 数据行
            for machine in range(p.machines_per_stage[stage]):
                ws_energy.cell(row, 1, f"M{stage+1},{machine+1}").border = thin_border
                col = 2
                for speed in range(p.n_speed_levels):
                    val = p.get_processing_power(stage, machine, speed)
                    cell = ws_energy.cell(row, col, round(val, 2))
                    cell.border = thin_border
                    col += 1
                ws_energy.cell(row, col, round(p.get_setup_power(stage, machine), 2)).border = thin_border
                ws_energy.cell(row, col + 1, round(p.get_idle_power(stage, machine), 2)).border = thin_border
                row += 1
            row += 1
        
        # 运输和辅助功率
        ws_energy.cell(row, 1, "运输功率 (kW)").font = Font(bold=True)
        ws_energy.cell(row, 2, p.transport_power)
        row += 1
        ws_energy.cell(row, 1, "辅助功率 (kW)").font = Font(bold=True)
        ws_energy.cell(row, 2, p.aux_power)
        
        # ===== 4. 工人参数 Sheet =====
        ws_worker = wb.create_sheet("工人参数")
        
        # 表头
        headers = ["技能等级", "工资(元/班次)", "可用人数"]
        for col, h in enumerate(headers, 1):
            cell = ws_worker.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 数据行
        for skill in range(p.n_skill_levels):
            ws_worker.cell(skill + 2, 1, chr(65 + skill)).border = thin_border
            ws_worker.cell(skill + 2, 2, float(p.skill_wages[skill])).border = thin_border
            ws_worker.cell(skill + 2, 3, int(p.workers_available[skill])).border = thin_border
        
        # 保存文件
        try:
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"数据已成功导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"保存文件时出错:\n{str(e)}")
    
    def create_processing_time_tab(self):
        """创建加工时间标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        layout.addWidget(QLabel("加工时间 (分钟): 每个工件在每个阶段、每台机器、每种速度下的加工时间"))
        
        # 选择器
        selector_layout = QHBoxLayout()
        selector_layout.addWidget(QLabel("选择阶段:"))
        self.proc_stage_combo = QComboBox()
        self.proc_stage_combo.currentIndexChanged.connect(self.update_processing_table)
        selector_layout.addWidget(self.proc_stage_combo)
        
        selector_layout.addWidget(QLabel("选择机器:"))
        self.proc_machine_combo = QComboBox()
        self.proc_machine_combo.currentIndexChanged.connect(self.update_processing_table)
        selector_layout.addWidget(self.proc_machine_combo)
        selector_layout.addStretch()
        layout.addLayout(selector_layout)
        
        # 表格: 行=工件, 列=速度等级
        self.proc_table = QTableWidget()
        layout.addWidget(self.proc_table)
        
        self.tabs.addTab(widget, "⏱️ 加工时间")
    
    def create_setup_time_tab(self):
        """创建换模时间标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        layout.addWidget(QLabel("换模时间 (分钟): 从工件A切换到工件B所需的准备时间"))
        
        selector_layout = QHBoxLayout()
        selector_layout.addWidget(QLabel("选择阶段:"))
        self.setup_stage_combo = QComboBox()
        self.setup_stage_combo.currentIndexChanged.connect(self.update_setup_table)
        selector_layout.addWidget(self.setup_stage_combo)
        
        selector_layout.addWidget(QLabel("选择机器:"))
        self.setup_machine_combo = QComboBox()
        self.setup_machine_combo.currentIndexChanged.connect(self.update_setup_table)
        selector_layout.addWidget(self.setup_machine_combo)
        selector_layout.addStretch()
        layout.addLayout(selector_layout)
        
        # 表格: 行=前工件, 列=后工件
        self.setup_table = QTableWidget()
        layout.addWidget(self.setup_table)
        
        self.tabs.addTab(widget, "🔧 换模时间")
    
    def create_energy_tab(self):
        """创建能耗参数标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 固定参数
        fixed_group = QGroupBox("固定能耗参数")
        fixed_layout = QGridLayout(fixed_group)
        
        fixed_layout.addWidget(QLabel("运输功率 (kW):"), 0, 0)
        self.transport_power_spin = QDoubleSpinBox()
        self.transport_power_spin.setRange(0.1, 10.0)
        self.transport_power_spin.setValue(0.5)
        self.transport_power_spin.setDecimals(2)
        fixed_layout.addWidget(self.transport_power_spin, 0, 1)
        
        fixed_layout.addWidget(QLabel("辅助功率 (kW):"), 0, 2)
        self.aux_power_spin = QDoubleSpinBox()
        self.aux_power_spin.setRange(0.1, 20.0)
        self.aux_power_spin.setValue(1.0)
        self.aux_power_spin.setDecimals(2)
        fixed_layout.addWidget(self.aux_power_spin, 0, 3)
        
        layout.addWidget(fixed_group)
        
        # 机器相关能耗
        layout.addWidget(QLabel("机器能耗 (kW): 加工功率(按速度)、换模功率、空闲功率"))
        
        selector_layout = QHBoxLayout()
        selector_layout.addWidget(QLabel("选择阶段:"))
        self.energy_stage_combo = QComboBox()
        self.energy_stage_combo.currentIndexChanged.connect(self.update_energy_table)
        selector_layout.addWidget(self.energy_stage_combo)
        selector_layout.addStretch()
        layout.addLayout(selector_layout)
        
        # 表格: 行=机器, 列=加工功率(各速度)+换模功率+空闲功率
        self.energy_table = QTableWidget()
        layout.addWidget(self.energy_table)
        
        self.tabs.addTab(widget, "⚡ 能耗参数")
    
    def create_worker_tab(self):
        """创建工人参数标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        layout.addWidget(QLabel("工人技能与工资: 每个技能等级的标准工期工资和可用人数"))
        
        self.worker_table = QTableWidget()
        layout.addWidget(self.worker_table)
        
        self.tabs.addTab(widget, "👷 工人参数")
    
    def load_problem_data(self):
        """加载问题数据到表格"""
        p = self.problem
        
        # 初始化选择器
        self.proc_stage_combo.clear()
        self.setup_stage_combo.clear()
        self.energy_stage_combo.clear()
        for s in range(p.n_stages):
            self.proc_stage_combo.addItem(f"阶段 {s+1}")
            self.setup_stage_combo.addItem(f"阶段 {s+1}")
            self.energy_stage_combo.addItem(f"阶段 {s+1}")
        
        self.update_machine_combos()
        
        # 固定参数
        self.transport_power_spin.setValue(p.transport_power)
        self.aux_power_spin.setValue(p.aux_power)
        
        # 工人表格
        self.worker_table.setRowCount(p.n_skill_levels)
        self.worker_table.setColumnCount(3)
        self.worker_table.setHorizontalHeaderLabels(["技能等级", "工资(元/班次)", "可用人数"])
        for skill in range(p.n_skill_levels):
            self.worker_table.setItem(skill, 0, QTableWidgetItem(chr(65 + skill)))  # A, B, C...
            
            wage_item = QTableWidgetItem(f"{p.skill_wages[skill]:.1f}")
            self.worker_table.setItem(skill, 1, wage_item)
            
            count_item = QTableWidgetItem(str(int(p.workers_available[skill])))
            self.worker_table.setItem(skill, 2, count_item)
        
        self.worker_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # 触发表格更新
        self.update_processing_table()
        self.update_setup_table()
        self.update_energy_table()
    
    def update_machine_combos(self):
        """更新机器选择器"""
        p = self.problem
        
        stage = self.proc_stage_combo.currentIndex()
        if stage >= 0 and stage < p.n_stages:
            self.proc_machine_combo.clear()
            for m in range(p.machines_per_stage[stage]):
                self.proc_machine_combo.addItem(f"M{stage+1},{m+1}")
        
        stage = self.setup_stage_combo.currentIndex()
        if stage >= 0 and stage < p.n_stages:
            self.setup_machine_combo.clear()
            for m in range(p.machines_per_stage[stage]):
                self.setup_machine_combo.addItem(f"M{stage+1},{m+1}")
    
    def update_processing_table(self):
        """更新加工时间表格"""
        if self.problem is None:
            return
        
        p = self.problem
        stage = self.proc_stage_combo.currentIndex()
        
        # 更新机器选择器
        self.proc_machine_combo.blockSignals(True)
        current_machine = self.proc_machine_combo.currentIndex()
        self.proc_machine_combo.clear()
        if stage >= 0 and stage < p.n_stages:
            for m in range(p.machines_per_stage[stage]):
                self.proc_machine_combo.addItem(f"M{stage+1},{m+1}")
            if current_machine >= 0 and current_machine < p.machines_per_stage[stage]:
                self.proc_machine_combo.setCurrentIndex(current_machine)
        self.proc_machine_combo.blockSignals(False)
        
        machine = self.proc_machine_combo.currentIndex()
        
        if stage < 0 or machine < 0:
            return
        
        self.proc_table.setRowCount(p.n_jobs)
        self.proc_table.setColumnCount(p.n_speed_levels)
        
        headers = [f"速度{s+1}" for s in range(p.n_speed_levels)]
        self.proc_table.setHorizontalHeaderLabels(headers)
        self.proc_table.setVerticalHeaderLabels([f"工件{j+1}" for j in range(p.n_jobs)])
        
        for job in range(p.n_jobs):
            for speed in range(p.n_speed_levels):
                val = p.processing_time[job, stage, machine, speed]
                item = QTableWidgetItem(f"{val:.0f}")
                self.proc_table.setItem(job, speed, item)
        
        self.proc_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    def update_setup_table(self):
        """更新换模时间表格"""
        if self.problem is None:
            return
        
        p = self.problem
        stage = self.setup_stage_combo.currentIndex()
        
        # 更新机器选择器
        self.setup_machine_combo.blockSignals(True)
        current_machine = self.setup_machine_combo.currentIndex()
        self.setup_machine_combo.clear()
        if stage >= 0 and stage < p.n_stages:
            for m in range(p.machines_per_stage[stage]):
                self.setup_machine_combo.addItem(f"M{stage+1},{m+1}")
            if current_machine >= 0 and current_machine < p.machines_per_stage[stage]:
                self.setup_machine_combo.setCurrentIndex(current_machine)
        self.setup_machine_combo.blockSignals(False)
        
        machine = self.setup_machine_combo.currentIndex()
        
        if stage < 0 or machine < 0:
            return
        
        self.setup_table.setRowCount(p.n_jobs)
        self.setup_table.setColumnCount(p.n_jobs)
        
        headers = [f"→工件{j+1}" for j in range(p.n_jobs)]
        self.setup_table.setHorizontalHeaderLabels(headers)
        self.setup_table.setVerticalHeaderLabels([f"工件{j+1}→" for j in range(p.n_jobs)])
        
        if p.setup_time is not None:
            for j1 in range(p.n_jobs):
                for j2 in range(p.n_jobs):
                    val = p.setup_time[stage, machine, j1, j2]
                    item = QTableWidgetItem(f"{val:.0f}")
                    self.setup_table.setItem(j1, j2, item)
        
        self.setup_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    def update_energy_table(self):
        """更新能耗表格"""
        if self.problem is None:
            return
        
        p = self.problem
        stage = self.energy_stage_combo.currentIndex()
        
        if stage < 0:
            return
        
        n_machines = p.machines_per_stage[stage]
        n_speeds = p.n_speed_levels
        
        self.energy_table.setRowCount(n_machines)
        self.energy_table.setColumnCount(n_speeds + 2)  # 速度 + 换模 + 空闲
        
        headers = [f"加工(速度{s+1})" for s in range(n_speeds)] + ["换模功率", "空闲功率"]
        self.energy_table.setHorizontalHeaderLabels(headers)
        self.energy_table.setVerticalHeaderLabels([f"M{stage+1},{m+1}" for m in range(n_machines)])
        
        for m in range(n_machines):
            for s in range(n_speeds):
                val = p.get_processing_power(stage, m, s)
                item = QTableWidgetItem(f"{val:.2f}")
                self.energy_table.setItem(m, s, item)
            
            setup_val = p.get_setup_power(stage, m)
            self.energy_table.setItem(m, n_speeds, QTableWidgetItem(f"{setup_val:.2f}"))
            
            idle_val = p.get_idle_power(stage, m)
            self.energy_table.setItem(m, n_speeds + 1, QTableWidgetItem(f"{idle_val:.2f}"))
        
        self.energy_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    def random_fill(self):
        """随机填充数据"""
        # 使用当前参数生成随机问题
        if self.params:
            machines_per_stage = [self.params.get('machines_per_stage', 2)] * self.params.get('n_stages', 3)
            self.problem = SchedulingProblem.generate_random(
                n_jobs=self.params.get('n_jobs', 5),
                n_stages=self.params.get('n_stages', 3),
                machines_per_stage=machines_per_stage,
                n_speed_levels=self.params.get('n_speed_levels', 3),
                n_skill_levels=self.params.get('n_skill_levels', 3),
                seed=self.params.get('seed', 42)
            )
            self.load_problem_data()
            QMessageBox.information(self, "提示", "已随机生成数据!")
    
    def get_problem(self) -> SchedulingProblem:
        """从表格读取并返回修改后的问题实例"""
        if self.problem is None:
            return None
        
        p = self.problem
        
        # 读取固定参数
        p.transport_power = self.transport_power_spin.value()
        p.aux_power = self.aux_power_spin.value()
        
        # 读取工人参数
        for skill in range(p.n_skill_levels):
            wage_item = self.worker_table.item(skill, 1)
            if wage_item:
                try:
                    p.skill_wages[skill] = float(wage_item.text())
                except:
                    pass
            
            count_item = self.worker_table.item(skill, 2)
            if count_item:
                try:
                    p.workers_available[skill] = int(count_item.text())
                except:
                    pass
        
        # 注意: 加工时间、换模时间、能耗表格的读取需要遍历所有阶段/机器
        # 这里简化处理，只读取当前显示的数据
        # 完整实现需要在切换选择器时保存之前的数据
        
        return p


class MainApp(QMainWindow):
    """
    主窗口 - 参数输入界面
    """
    
    def __init__(self):
        super().__init__()
        self.problem = None
        self.initial_solution = None
        self.setup_ui()
        self.setStyleSheet(MAIN_STYLE)
    
    def setup_ui(self):
        """初始化UI - 卡片式布局"""
        self.setWindowTitle("多目标调度优化系统 - 参数设置")
        self.setMinimumSize(900, 720)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(25, 20, 25, 20)
        
        # ===== 标题 =====
        title_label = QLabel("🏭 三目标混合流水车间调度优化系统")
        title_label.setFont(QFont("Microsoft YaHei", 20, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"""
            color: {COLORS['text_light']};
            padding: 20px;
            background-color: {COLORS['dark_card']};
            border-radius: 16px;
            font-size: 20px;
        """)
        main_layout.addWidget(title_label)
        
        # ===== 田口设计按钮 和 算法对比按钮 =====
        taguchi_btn_layout = QHBoxLayout()
        taguchi_btn_layout.addStretch()
        
        self.taguchi_btn = QPushButton("🔬 田口设计 (参数调优)")
        self.taguchi_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['secondary']};
                color: {COLORS['text_light']};
                border: none;
                padding: 14px 28px;
                border-radius: 12px;
                font-size: 15px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #2980b9;
            }}
        """)
        self.taguchi_btn.setToolTip("使用田口实验法 L16(4⁴) 正交表进行算法参数调优")
        self.taguchi_btn.clicked.connect(self.on_taguchi_design)
        taguchi_btn_layout.addWidget(self.taguchi_btn)
        
        # 算法对比试验按钮
        self.comparison_btn = QPushButton("📊 算法对比试验")
        self.comparison_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #8e44ad;
                color: {COLORS['text_light']};
                border: none;
                padding: 14px 28px;
                border-radius: 12px;
                font-size: 15px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #7d3c98;
            }}
        """)
        self.comparison_btn.setToolTip("对比多种优化算法的性能（IGD/HV/GD指标）")
        self.comparison_btn.clicked.connect(self.on_algorithm_comparison)
        taguchi_btn_layout.addWidget(self.comparison_btn)
        
        taguchi_btn_layout.addStretch()
        main_layout.addLayout(taguchi_btn_layout)
        
        # 创建滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(12)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        
        # ===== 1. 数据输入模式 (居中顶部) =====
        self.create_mode_group(scroll_layout)
        
        # ===== 2-4. 三个参数模块横向排列 (各占一半) =====
        params_row = QHBoxLayout()
        params_row.setSpacing(15)
        
        # 左侧：问题参数 (占一半)
        problem_widget = QWidget()
        problem_layout = QVBoxLayout(problem_widget)
        problem_layout.setContentsMargins(0, 0, 0, 0)
        self.create_problem_group(problem_layout)
        params_row.addWidget(problem_widget, stretch=1)
        
        # 右侧：算法参数 + 目标权重 垂直排列 (占一半)
        right_column = QVBoxLayout()
        right_column.setSpacing(12)
        
        self.create_algorithm_group(right_column)
        self.create_weights_group(right_column)
        right_column.addStretch()
        
        right_widget = QWidget()
        right_widget.setLayout(right_column)
        params_row.addWidget(right_widget, stretch=1)
        
        scroll_layout.addLayout(params_row)
        scroll_layout.addStretch()
        
        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll)
        
        # ===== 底部按钮区域 (居中) =====
        param_btn_layout = QHBoxLayout()
        param_btn_layout.setSpacing(15)
        param_btn_layout.addStretch()  # 左侧弹性
        
        self.lock_btn = QPushButton("🔒 确认参数")
        self.lock_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['success']};
                color: {COLORS['text_light']};
                border: none;
                padding: 18px 48px;
                border-radius: 14px;
                font-weight: bold;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: #219a52;
            }}
        """)
        self.lock_btn.clicked.connect(self.on_lock_params)
        param_btn_layout.addWidget(self.lock_btn)
        
        self.unlock_btn = QPushButton("🔓 修改参数")
        self.unlock_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['warning']};
                color: {COLORS['text_dark']};
                border: none;
                padding: 18px 48px;
                border-radius: 14px;
                font-weight: bold;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: #e09612;
            }}
        """)
        self.unlock_btn.clicked.connect(self.on_unlock_params)
        self.unlock_btn.setEnabled(False)
        param_btn_layout.addWidget(self.unlock_btn)
        
        # 参数状态标签
        self.param_status_label = QLabel("⚠️ 参数未确认")
        self.param_status_label.setStyleSheet("color: #dc3545; font-weight: bold; background-color: transparent; margin-left: 15px;")
        param_btn_layout.addWidget(self.param_status_label)
        
        param_btn_layout.addStretch()  # 右侧弹性
        
        main_layout.addLayout(param_btn_layout)
        
        # 生成编码按钮
        self.confirm_btn = QPushButton("✅ 确认参数并生成编码")
        self.confirm_btn.setStyleSheet(PRIMARY_BUTTON_STYLE)
        self.confirm_btn.clicked.connect(self.on_confirm)
        self.confirm_btn.setEnabled(False)
        main_layout.addWidget(self.confirm_btn)
        
        # 参数锁定状态
        self.params_locked = False
    
    def on_lock_params(self):
        """锁定参数"""
        self.params_locked = True
        self._set_params_enabled(False)
        self.lock_btn.setEnabled(False)
        self.unlock_btn.setEnabled(True)
        self.confirm_btn.setEnabled(True)
        self.param_status_label.setText("✅ 参数已确认")
        self.param_status_label.setStyleSheet("color: #28a745; font-weight: bold;")
        QMessageBox.information(self, "提示", "参数已锁定！\n可以点击\"确认参数并生成编码\"继续。")
    
    def on_unlock_params(self):
        """解锁参数"""
        self.params_locked = False
        self._set_params_enabled(True)
        self.lock_btn.setEnabled(True)
        self.unlock_btn.setEnabled(False)
        self.confirm_btn.setEnabled(False)
        self.param_status_label.setText("⚠️ 参数未确认")
        self.param_status_label.setStyleSheet("color: #dc3545; font-weight: bold;")
        # 清除已生成的问题实例
        self.problem = None
    
    def _set_params_enabled(self, enabled: bool):
        """设置所有参数输入控件的启用/禁用状态"""
        # 问题参数
        self.n_jobs_spin.setEnabled(enabled)
        self.n_stages_spin.setEnabled(enabled)
        self.machines_spin.setEnabled(enabled)
        self.speed_levels_spin.setEnabled(enabled)
        self.skill_levels_spin.setEnabled(enabled)
        self.seed_spin.setEnabled(enabled)
        
        # 算法参数
        self.pop_size_spin.setEnabled(enabled)
        self.n_gen_spin.setEnabled(enabled)
        self.crossover_spin.setEnabled(enabled)
        self.mutation_spin.setEnabled(enabled)
        self.init_temp_spin.setEnabled(enabled)
        self.cooling_spin.setEnabled(enabled)
        
        # 目标权重
        self.w1_spin.setEnabled(enabled)
        self.w2_spin.setEnabled(enabled)
        self.w3_spin.setEnabled(enabled)
        
        # 数据模式
        self.auto_radio.setEnabled(enabled)
        self.manual_radio.setEnabled(enabled)
        self.auto_confirm_btn.setEnabled(enabled)
        self.manual_input_btn.setEnabled(enabled)
        
        # 工人数量
        for spin in self.worker_count_spins:
            spin.setEnabled(enabled)
    
    def create_mode_group(self, parent_layout):
        """创建数据模式选择组 - 内容居中"""
        group = QGroupBox("📊 数据输入模式")
        layout = QHBoxLayout(group)
        layout.setAlignment(Qt.AlignCenter)  # 内容居中
        layout.setSpacing(20)
        
        self.mode_group = QButtonGroup(self)
        
        self.auto_radio = QRadioButton("🎲 自动生成数据")
        self.auto_radio.setChecked(True)
        self.auto_radio.setStyleSheet(f"font-size: 15px; font-weight: bold; padding: 8px; color: {COLORS['text_dark']};")
        self.manual_radio = QRadioButton("✏️ 手动输入数据")
        self.manual_radio.setStyleSheet(f"font-size: 15px; font-weight: bold; padding: 8px; color: {COLORS['text_dark']};")
        
        self.mode_group.addButton(self.auto_radio, 0)
        self.mode_group.addButton(self.manual_radio, 1)
        
        layout.addStretch()  # 左侧弹性
        layout.addWidget(self.auto_radio)
        
        # 自动生成确认按钮
        self.auto_confirm_btn = QPushButton("📋 查看/编辑数据")
        self.auto_confirm_btn.setStyleSheet(SECONDARY_BUTTON_STYLE)
        self.auto_confirm_btn.clicked.connect(self.on_view_auto_data)
        layout.addWidget(self.auto_confirm_btn)
        
        layout.addWidget(self.manual_radio)
        
        # 手动输入按钮
        self.manual_input_btn = QPushButton("📝 输入数据")
        self.manual_input_btn.setStyleSheet(SECONDARY_BUTTON_STYLE)
        self.manual_input_btn.clicked.connect(self.on_manual_input)
        layout.addWidget(self.manual_input_btn)
        
        layout.addStretch()  # 右侧弹性
        
        parent_layout.addWidget(group)
    
    def create_problem_group(self, parent_layout):
        """创建问题参数组"""
        group = QGroupBox("🔧 问题参数")
        layout = QGridLayout(group)
        layout.setSpacing(10)
        
        row = 0
        
        layout.addWidget(QLabel("工件数量:"), row, 0)
        self.n_jobs_spin = QSpinBox()
        self.n_jobs_spin.setRange(2, 100)
        self.n_jobs_spin.setValue(5)
        layout.addWidget(self.n_jobs_spin, row, 1)
        
        layout.addWidget(QLabel("阶段数量:"), row, 2)
        self.n_stages_spin = QSpinBox()
        self.n_stages_spin.setRange(2, 20)
        self.n_stages_spin.setValue(3)
        layout.addWidget(self.n_stages_spin, row, 3)
        
        row += 1
        
        layout.addWidget(QLabel("每阶段机器数:"), row, 0)
        self.machines_spin = QSpinBox()
        self.machines_spin.setRange(1, 10)
        self.machines_spin.setValue(2)
        layout.addWidget(self.machines_spin, row, 1)
        
        layout.addWidget(QLabel("速度等级数:"), row, 2)
        self.speed_levels_spin = QSpinBox()
        self.speed_levels_spin.setRange(1, 5)
        self.speed_levels_spin.setValue(3)
        layout.addWidget(self.speed_levels_spin, row, 3)
        
        row += 1
        
        layout.addWidget(QLabel("技能等级数:"), row, 0)
        self.skill_levels_spin = QSpinBox()
        self.skill_levels_spin.setRange(1, 5)
        self.skill_levels_spin.setValue(3)
        self.skill_levels_spin.valueChanged.connect(self.on_skill_levels_changed)
        layout.addWidget(self.skill_levels_spin, row, 1)
        
        layout.addWidget(QLabel("随机种子:"), row, 2)
        self.seed_spin = QSpinBox()
        self.seed_spin.setRange(0, 99999)
        self.seed_spin.setValue(42)
        layout.addWidget(self.seed_spin, row, 3)
        
        row += 1
        
        # 各技能等级工人数量
        layout.addWidget(QLabel("各技能等级工人数量:"), row, 0, 1, 4)
        row += 1
        
        # 创建工人数量输入容器
        self.worker_count_widget = QWidget()
        self.worker_count_layout = QHBoxLayout(self.worker_count_widget)
        self.worker_count_layout.setContentsMargins(0, 0, 0, 0)
        self.worker_count_spins = []
        self.worker_id_labels = []  # 显示工人编号的标签
        
        # 默认创建3个技能等级的工人数量输入
        self._create_worker_count_inputs(3)
        
        layout.addWidget(self.worker_count_widget, row, 0, 1, 4)
        
        parent_layout.addWidget(group)
    
    def _create_worker_count_inputs(self, n_skill_levels: int):
        """创建工人数量输入控件"""
        # 清除旧的控件
        for spin in self.worker_count_spins:
            spin.setParent(None)
        self.worker_count_spins.clear()
        
        for label in self.worker_id_labels:
            label.setParent(None)
        self.worker_id_labels.clear()
        
        # 清除布局中的所有项
        while self.worker_count_layout.count():
            item = self.worker_count_layout.takeAt(0)
            if item.widget():
                item.widget().setParent(None)
        
        # 创建新的输入控件
        default_counts = [5, 4, 3, 2, 2]  # 默认各等级工人数
        skill_letters = ['A', 'B', 'C', 'D', 'E']  # 技能等级字母
        
        for i in range(n_skill_levels):
            skill_letter = skill_letters[i] if i < len(skill_letters) else chr(ord('A') + i)
            
            # 等级标签
            label = QLabel(f"等级{skill_letter}:")
            self.worker_count_layout.addWidget(label)
            
            # 数量输入
            spin = QSpinBox()
            spin.setRange(1, 20)
            spin.setValue(default_counts[i] if i < len(default_counts) else 2)
            spin.valueChanged.connect(self._update_worker_labels)
            self.worker_count_spins.append(spin)
            self.worker_count_layout.addWidget(spin)
            
            # 工人编号标签
            id_label = QLabel()
            id_label.setStyleSheet("color: #007bff; font-size: 11px;")
            self.worker_id_labels.append(id_label)
            self.worker_count_layout.addWidget(id_label)
        
        self.worker_count_layout.addStretch()
        
        # 初始化工人编号显示
        self._update_worker_labels()
    
    def _update_worker_labels(self):
        """更新工人编号标签显示"""
        skill_letters = ['A', 'B', 'C', 'D', 'E']
        
        for i, (spin, id_label) in enumerate(zip(self.worker_count_spins, self.worker_id_labels)):
            count = spin.value()
            skill_letter = skill_letters[i] if i < len(skill_letters) else chr(ord('A') + i)
            
            # 生成工人编号列表
            worker_ids = [f"{skill_letter}{j+1}" for j in range(count)]
            full_text = ', '.join(worker_ids)
            
            # 如果超过4个工人，显示省略形式并设置tooltip
            if count > 4:
                short_text = f"{skill_letter}1~{skill_letter}{count}"
                id_label.setText(f"({short_text})")
                id_label.setToolTip(f"所有工人: {full_text}")
                id_label.setStyleSheet("color: #007bff; font-size: 11px; text-decoration: underline;")
            else:
                id_label.setText(f"({full_text})")
                id_label.setToolTip("")
                id_label.setStyleSheet("color: #007bff; font-size: 11px;")

    
    def on_skill_levels_changed(self, value: int):
        """技能等级数变化时重新创建工人数量输入"""
        self._create_worker_count_inputs(value)
    
    def create_algorithm_group(self, parent_layout):
        """创建算法参数组 - 移除MOSA/VNS迭代次数"""
        group = QGroupBox("⚙️ 算法参数")
        layout = QGridLayout(group)
        layout.setSpacing(10)
        
        row = 0
        
        # NSGA-II 参数
        layout.addWidget(QLabel("NSGA-II 种群大小:"), row, 0)
        self.pop_size_spin = QSpinBox()
        self.pop_size_spin.setRange(10, 500)
        self.pop_size_spin.setValue(50)
        layout.addWidget(self.pop_size_spin, row, 1)
        
        layout.addWidget(QLabel("进化代数:"), row, 2)
        self.n_gen_spin = QSpinBox()
        self.n_gen_spin.setRange(10, 1000)
        self.n_gen_spin.setValue(100)
        layout.addWidget(self.n_gen_spin, row, 3)
        
        row += 1
        
        layout.addWidget(QLabel("Crossover Rate:"), row, 0)
        self.crossover_spin = QDoubleSpinBox()
        self.crossover_spin.setRange(0.0, 1.0)
        self.crossover_spin.setSingleStep(0.1)
        self.crossover_spin.setValue(0.9)
        layout.addWidget(self.crossover_spin, row, 1)
        
        layout.addWidget(QLabel("Mutation Rate:"), row, 2)
        self.mutation_spin = QDoubleSpinBox()
        self.mutation_spin.setRange(0.0, 1.0)
        self.mutation_spin.setSingleStep(0.05)
        self.mutation_spin.setValue(0.1)
        layout.addWidget(self.mutation_spin, row, 3)
        
        row += 1
        
        # MOSA 参数 (只保留温度相关)
        layout.addWidget(QLabel("MOSA Initial Temp:"), row, 0)
        self.init_temp_spin = QDoubleSpinBox()
        self.init_temp_spin.setRange(1.0, 1000.0)
        self.init_temp_spin.setValue(100.0)
        layout.addWidget(self.init_temp_spin, row, 1)
        
        layout.addWidget(QLabel("Cooling Rate:"), row, 2)
        self.cooling_spin = QDoubleSpinBox()
        self.cooling_spin.setRange(0.8, 0.999)
        self.cooling_spin.setSingleStep(0.01)
        self.cooling_spin.setDecimals(3)
        self.cooling_spin.setValue(0.95)
        layout.addWidget(self.cooling_spin, row, 3)
        
        row += 1
        
        # 提示信息
        hint = QLabel("提示: VNS/MOSA 迭代次数 = NSGA-II 输出的 Pareto 解数量")
        hint.setStyleSheet("color: #7f8c8d; font-style: italic;")
        layout.addWidget(hint, row, 0, 1, 4)
        
        parent_layout.addWidget(group)
    
    def create_weights_group(self, parent_layout):
        """创建目标权重组"""
        group = QGroupBox("⚖️ 目标权重")
        layout = QGridLayout(group)
        layout.setSpacing(10)
        
        layout.addWidget(QLabel("F1 (Makespan) 权重:"), 0, 0)
        self.w1_spin = QDoubleSpinBox()
        self.w1_spin.setRange(0.0, 10.0)
        self.w1_spin.setSingleStep(0.1)
        self.w1_spin.setValue(1.0)
        layout.addWidget(self.w1_spin, 0, 1)
        
        layout.addWidget(QLabel("F2 (Labor Cost) 权重:"), 0, 2)
        self.w2_spin = QDoubleSpinBox()
        self.w2_spin.setRange(0.0, 10.0)
        self.w2_spin.setSingleStep(0.1)
        self.w2_spin.setValue(1.0)
        layout.addWidget(self.w2_spin, 0, 3)
        
        layout.addWidget(QLabel("F3 (Energy) 权重:"), 1, 0)
        self.w3_spin = QDoubleSpinBox()
        self.w3_spin.setRange(0.0, 10.0)
        self.w3_spin.setSingleStep(0.1)
        self.w3_spin.setValue(1.0)
        layout.addWidget(self.w3_spin, 1, 1)
        
        parent_layout.addWidget(group)
    
    def get_parameters(self) -> dict:
        """获取所有参数"""
        w1 = self.w1_spin.value()
        w2 = self.w2_spin.value()
        w3 = self.w3_spin.value()
        total = w1 + w2 + w3
        if total > 0:
            weights = (w1/total, w2/total, w3/total)
        else:
            weights = (1/3, 1/3, 1/3)
        
        # 获取各技能等级工人数量
        workers_per_skill = [spin.value() for spin in self.worker_count_spins]
        
        return {
            'auto_mode': self.auto_radio.isChecked(),
            'n_jobs': self.n_jobs_spin.value(),
            'n_stages': self.n_stages_spin.value(),
            'machines_per_stage': self.machines_spin.value(),
            'n_speed_levels': self.speed_levels_spin.value(),
            'n_skill_levels': self.skill_levels_spin.value(),
            'workers_per_skill': workers_per_skill,
            'seed': self.seed_spin.value(),
            'pop_size': self.pop_size_spin.value(),
            'n_generations': self.n_gen_spin.value(),
            'crossover_prob': self.crossover_spin.value(),
            'mutation_prob': self.mutation_spin.value(),
            'initial_temp': self.init_temp_spin.value(),
            'cooling_rate': self.cooling_spin.value(),
            'final_temp': 1.0,
            'weights': weights
        }
    
    def on_view_auto_data(self):
        """查看/编辑自动生成的数据"""
        params = self.get_parameters()
        
        # 生成问题
        machines_per_stage = [params['machines_per_stage']] * params['n_stages']
        self.problem = SchedulingProblem.generate_random(
            n_jobs=params['n_jobs'],
            n_stages=params['n_stages'],
            machines_per_stage=machines_per_stage,
            n_speed_levels=params['n_speed_levels'],
            n_skill_levels=params['n_skill_levels'],
            seed=params['seed']
        )
        
        # 使用主页面输入的工人数量覆盖随机生成的值
        import numpy as np
        workers_per_skill = params.get('workers_per_skill', [])
        if workers_per_skill:
            self.problem.workers_available = np.array(workers_per_skill[:params['n_skill_levels']])
        
        dialog = DataEditorDialog(self, self.problem, is_manual=False, params=params)
        if dialog.exec_() == QDialog.Accepted:
            self.problem = dialog.get_problem()
            QMessageBox.information(self, "提示", "数据已确认!")
    
    def on_manual_input(self):
        """手动输入数据"""
        params = self.get_parameters()
        
        # 创建空问题结构，用于手动输入
        machines_per_stage = [params['machines_per_stage']] * params['n_stages']
        self.problem = SchedulingProblem.generate_random(
            n_jobs=params['n_jobs'],
            n_stages=params['n_stages'],
            machines_per_stage=machines_per_stage,
            n_speed_levels=params['n_speed_levels'],
            n_skill_levels=params['n_skill_levels'],
            seed=params['seed']
        )
        
        dialog = DataEditorDialog(self, self.problem, is_manual=True, params=params)
        if dialog.exec_() == QDialog.Accepted:
            self.problem = dialog.get_problem()
            QMessageBox.information(self, "提示", "数据已保存!")
    
    def on_taguchi_design(self):
        """打开田口实验设计窗口"""
        from ui.taguchi_window import TaguchiWindow
        
        self.taguchi_window = TaguchiWindow(self)
        self.taguchi_window.show()
    
    def on_algorithm_comparison(self):
        """打开算法对比试验窗口"""
        from ui.algorithm_comparison_window import AlgorithmComparisonWindow
        
        self.comparison_window = AlgorithmComparisonWindow(self)
        self.comparison_window.show()
    
    def on_confirm(self):
        """确认参数并生成编码"""
        params = self.get_parameters()
        
        try:
            # 始终根据当前参数重新生成问题（确保使用最新设置）
            machines_per_stage = [params['machines_per_stage']] * params['n_stages']
            self.problem = SchedulingProblem.generate_random(
                n_jobs=params['n_jobs'],
                n_stages=params['n_stages'],
                machines_per_stage=machines_per_stage,
                n_speed_levels=params['n_speed_levels'],
                n_skill_levels=params['n_skill_levels'],
                seed=params['seed']
            )
            
            # 使用主页面输入的工人数量覆盖随机生成的值
            import numpy as np
            workers_per_skill = params.get('workers_per_skill', [])
            if workers_per_skill:
                self.problem.workers_available = np.array(workers_per_skill[:params['n_skill_levels']])
            
            # 生成初始解
            self.initial_solution = Solution.generate_random(self.problem, seed=params['seed'])
            
            # 打开矩阵编码窗口
            self.matrix_window = MatrixWindow(self.problem, self.initial_solution, params)
            self.matrix_window.show()
            
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "错误", f"生成数据时出错:\n{str(e)}\n{traceback.format_exc()}")


class MatrixWindow(QMainWindow):
    """
    矩阵编码窗口 - 显示M/Q/V/W四矩阵
    """
    
    def __init__(self, problem: SchedulingProblem, solution: Solution, params: dict):
        super().__init__()
        self.problem = problem
        self.solution = solution
        self.params = params
        self.setup_ui()
        self.setStyleSheet(MAIN_STYLE)
    
    def setup_ui(self):
        """初始化UI"""
        self.setWindowTitle("四矩阵编码 (M-Q-V-W)")
        self.setMinimumSize(900, 700)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # 标题
        title = QLabel("📋 初始解的四矩阵编码")
        title.setFont(QFont("Microsoft YaHei", 14, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title)
        
        # 问题摘要
        summary = QLabel(f"工件数: {self.problem.n_jobs} | 阶段数: {self.problem.n_stages} | "
                        f"机器配置: {self.problem.machines_per_stage}")
        summary.setAlignment(Qt.AlignCenter)
        summary.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        main_layout.addWidget(summary)
        
        # 红色加粗提示
        notice = QLabel("⚠️ 此界面仅展示一个随机初始解的编码格式，帮助您了解数据结构。\n真正的优化结果需要点击\"开始优化\"运行算法后才会得到。")
        notice.setAlignment(Qt.AlignCenter)
        notice.setWordWrap(True)
        notice.setStyleSheet("""
            color: #dc3545;
            font-weight: bold;
            font-size: 12px;
            background-color: #fff3cd;
            border: 1px solid #ffc107;
            border-radius: 4px;
            padding: 8px;
            margin: 5px 0;
        """)
        main_layout.addWidget(notice)
        
        # 创建标签页显示四个矩阵
        tabs = QTabWidget()
        
        # M矩阵
        m_widget = self.create_matrix_display("M - Machine Assignment", 
                                               self.solution.machine_assign,
                                               "M[i,j] = Machine ID for Job i at Stage j (0-based)")
        tabs.addTab(m_widget, "📍 Machine (M)")
        
        # Q矩阵
        q_widget = self.create_matrix_display("Q - Sequence Priority",
                                               self.solution.sequence_priority,
                                               "Q[i,j] = Priority key for Job i at Stage j (smaller = higher priority)")
        tabs.addTab(q_widget, "🔢 Priority (Q)")
        
        # V矩阵
        v_widget = self.create_matrix_display("V - Speed Level",
                                               self.solution.speed_level,
                                               "V[i,j] = Speed level for Job i at Stage j (0=Low, 1=Medium, 2=High)")
        tabs.addTab(v_widget, "⚡ Speed (V)")
        
        # W矩阵
        w_widget = self.create_matrix_display("W - Worker Skill",
                                               self.solution.worker_skill,
                                               "W[i,j] = Worker skill level for Job i at Stage j (0-based)")
        tabs.addTab(w_widget, "👷 Worker (W)")
        
        main_layout.addWidget(tabs)
        
        # 按钮区
        btn_layout = QHBoxLayout()
        
        back_btn = QPushButton("⬅️ 返回")
        back_btn.clicked.connect(self.close)
        btn_layout.addWidget(back_btn)
        
        btn_layout.addStretch()
        
        self.start_btn = QPushButton("🚀 开始优化")
        self.start_btn.setStyleSheet(PRIMARY_BUTTON_STYLE)
        self.start_btn.clicked.connect(self.on_start_optimization)
        btn_layout.addWidget(self.start_btn)
        
        main_layout.addLayout(btn_layout)
    
    def create_matrix_display(self, title: str, matrix, description: str, 
                               display_offset: int = 0) -> QWidget:
        """
        创建矩阵显示组件
        
        Args:
            display_offset: 显示偏移量，用于将0-based转为1-based显示
        """
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        desc_label = QLabel(description)
        desc_label.setStyleSheet("color: #7f8c8d; font-style: italic;")
        layout.addWidget(desc_label)
        
        # 使用表格显示矩阵
        if matrix is None:
            error_label = QLabel("Matrix data is None!")
            error_label.setStyleSheet("color: red;")
            layout.addWidget(error_label)
            return widget
        
        n_jobs, n_stages = matrix.shape
        
        table = QTableWidget()
        table.setRowCount(n_jobs)
        table.setColumnCount(n_stages)
        
        # 设置表头
        table.setHorizontalHeaderLabels([f"Stage {s+1}" for s in range(n_stages)])
        table.setVerticalHeaderLabels([f"Job {j+1}" for j in range(n_jobs)])
        
        # 填充数据 (加上偏移量显示)
        for job in range(n_jobs):
            for stage in range(n_stages):
                value = int(matrix[job, stage]) + display_offset
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                table.setItem(job, stage, item)
        
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        layout.addWidget(table)
        
        return widget
    
    def on_start_optimization(self):
        """开始优化"""
        from ui.result_window import ResultWindow
        
        self.result_window = ResultWindow(self.problem, self.params)
        self.result_window.show()
        self.result_window.start_optimization()


def main():
    """程序入口"""
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    window = MainApp()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
